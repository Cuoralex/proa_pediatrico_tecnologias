from __future__ import annotations

from pathlib import Path
import argparse
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def _is_empty_value(value: object) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _clean_sheet(worksheet) -> None:
    for row_idx in range(worksheet.max_row, 0, -1):
        if all(_is_empty_value(cell.value) for cell in worksheet[row_idx]):
            worksheet.delete_rows(row_idx)

    for col_idx in range(worksheet.max_column, 0, -1):
        column_cells = [worksheet.cell(row=row_idx, column=col_idx) for row_idx in range(1, worksheet.max_row + 1)]
        if all(_is_empty_value(cell.value) for cell in column_cells):
            worksheet.delete_cols(col_idx)


def clean_excel_file(input_path: str | Path, output_path: str | Path | None = None) -> Path:
    """
    Limpia un archivo `.xlsx` eliminando filas y columnas completamente vacías.

    Args:
        input_path: Ruta al archivo de entrada.
        output_path: Ruta de salida opcional. Si no se indica, se crea
            `<nombre>_cleaned.xlsx` en la misma carpeta del archivo de entrada.

    Returns:
        Ruta del archivo limpio generado.

    Raises:
        ValueError: Si el archivo no tiene extensión `.xlsx`.
        FileNotFoundError: Si el archivo de entrada no existe.
        RuntimeError: Si no se puede leer el archivo Excel.
    """
    source = Path(input_path)
    if source.suffix.lower() != ".xlsx":
        raise ValueError("Solo se soportan archivos .xlsx")
    if not source.exists():
        raise FileNotFoundError(f"No se encontró el archivo: {source}")

    destination = Path(output_path) if output_path else source.with_name(f"{source.stem}_cleaned.xlsx")

    try:
        workbook = load_workbook(source)
    except (InvalidFileException, BadZipFile, OSError) as exc:
        raise RuntimeError(f"No se pudo leer el archivo Excel: {source}") from exc
    for sheet in workbook.worksheets:
        _clean_sheet(sheet)
    workbook.save(destination)

    return destination


def main() -> None:
    parser = argparse.ArgumentParser(description="Limpia filas y columnas vacías de un archivo Excel (.xlsx).")
    parser.add_argument("input_path", help="Ruta del archivo de entrada")
    parser.add_argument("--output-path", dest="output_path", help="Ruta del archivo limpio de salida")
    args = parser.parse_args()

    result = clean_excel_file(args.input_path, args.output_path)
    print(f"Archivo limpio generado: {result}")


if __name__ == "__main__":
    main()
