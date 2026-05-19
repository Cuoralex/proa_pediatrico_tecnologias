import sys
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook, load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "src"))

from proa_pediatrico_tecnologias.clean_excel import clean_excel_file


class CleanExcelFileTests(unittest.TestCase):
    def test_removes_empty_rows_and_columns(self):
        with TemporaryDirectory() as tmp_dir:
            input_file = Path(tmp_dir) / "entrada.xlsx"

            wb = Workbook()
            ws = wb.active
            ws["A1"] = "Nombre"
            ws["B1"] = None
            ws["A2"] = "Ana"
            ws["B2"] = None
            ws["A3"] = None
            ws["B3"] = None
            wb.save(input_file)

            output_file = clean_excel_file(input_file)
            self.assertTrue(output_file.exists())
            self.assertEqual(output_file.name, "entrada_cleaned.xlsx")

            cleaned_wb = load_workbook(output_file)
            cleaned_ws = cleaned_wb.active

            self.assertEqual(cleaned_ws.max_row, 2)
            self.assertEqual(cleaned_ws.max_column, 1)
            self.assertEqual(cleaned_ws["A1"].value, "Nombre")
            self.assertEqual(cleaned_ws["A2"].value, "Ana")

    def test_rejects_non_xlsx_input(self):
        with self.assertRaises(ValueError):
            clean_excel_file("archivo.csv")
        with self.assertRaises(ValueError):
            clean_excel_file(Path("archivo.csv"))

    def test_raises_file_not_found_for_missing_file(self):
        with self.assertRaises(FileNotFoundError):
            clean_excel_file(Path("no-existe.xlsx"))

    def test_raises_runtime_error_for_invalid_xlsx_content(self):
        with TemporaryDirectory() as tmp_dir:
            invalid_xlsx = Path(tmp_dir) / "invalido.xlsx"
            invalid_xlsx.write_text("no es un archivo xlsx válido", encoding="utf-8")
            with self.assertRaisesRegex(RuntimeError, "No se pudo leer el archivo Excel"):
                clean_excel_file(invalid_xlsx)

    def test_rejects_same_input_and_output_path(self):
        with TemporaryDirectory() as tmp_dir:
            input_file = Path(tmp_dir) / "entrada.xlsx"
            wb = Workbook()
            wb.active["A1"] = "dato"
            wb.save(input_file)

            with self.assertRaises(ValueError):
                clean_excel_file(input_file, input_file)


if __name__ == "__main__":
    unittest.main()
